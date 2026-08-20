"""Parser měsíčního exportu vyúčtování sdílení (GoEnergy).

Formát zjištěný rozborem reálného souboru, zdokumentovaný v
`dev/step09/SHARING_IMPORT_FORMAT.md`. Čisté funkce bez vedlejších účinků
na disk/síť kromě samotného otevření zadaného souboru – testovatelné bez
readeru. `sharing_reader.py` výstup mapuje na `EnergyInterval`.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

# Sloupce intervalové tabulky (řádek 1 = hlavička).
COLUMNS = {
    "ean": "EAN",
    "day": "DEN",
    "hour": "HODINA",
    "minute": "MINUTA",
    "delivered_kwh": "Výroba kWh (dodávka do sítě)",
    "sold_kwh": "Výroba kWh (po odečtení sdílení)",
    "shared_kwh": "Sdílená energie (kWH)",
}

# Popisky měsíčního souhrnu ve sloupci K – řádky se dohledávají podle textu,
# ne podle pevného čísla řádku (odolnější vůči drobným změnám formátu).
SETTLEMENT_LABELS = {
    "delivered_kwh": "Celková dodávka do sítě",
    "shared_kwh": "Sdílená elektřina",
    "sold_kwh": "Elektřina prodaná GoEnergy",
    "deduction_czk": "Srážka z ceny",
    "invoice_czk": "Fakturace",
}
SETTLEMENT_LABEL_COLUMN = 11    # K
SETTLEMENT_AMOUNT_COLUMN = 12   # L – množství (MWh u výše uvedených řádků)
SETTLEMENT_TOTAL_COLUMN = 14    # N – výsledná cena (Kč) u srážky/fakturace
SETTLEMENT_SEARCH_ROWS = 30


def _load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as exc:                                  # pragma: no cover
        raise RuntimeError("knihovna 'openpyxl' není nainstalována – "
                           "je potřeba jen pro import vyúčtování sdílení") from exc
    return openpyxl.load_workbook(path, data_only=True)


def _header_columns(ws) -> dict[str, int]:
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    missing = [label for label in COLUMNS.values() if label not in headers]
    if missing:
        raise ValueError(f"neočekávaný formát exportu GoEnergy, chybí sloupce: {missing}")
    return {key: headers[label] for key, label in COLUMNS.items()}


def _parse_intervals(ws, columns: dict[str, int]) -> list[dict[str, Any]]:
    intervals: list[dict[str, Any]] = []
    ean = None
    for row in range(2, ws.max_row + 1):
        day_value = ws.cell(row=row, column=columns["day"]).value
        if day_value is None:
            continue
        ean = ws.cell(row=row, column=columns["ean"]).value or ean
        hour = int(ws.cell(row=row, column=columns["hour"]).value or 0)
        minute = int(ws.cell(row=row, column=columns["minute"]).value or 0)
        moment = datetime(day_value.year, day_value.month, day_value.day, hour, minute)
        intervals.append({
            "timestamp": moment,
            "ean": ean,
            "production_kwh": ws.cell(row=row, column=columns["delivered_kwh"]).value,
            "sold_kwh": ws.cell(row=row, column=columns["sold_kwh"]).value,
            "shared_kwh": ws.cell(row=row, column=columns["shared_kwh"]).value,
        })
    return intervals


def _find_label_row(ws, label: str) -> int | None:
    for row in range(1, SETTLEMENT_SEARCH_ROWS + 1):
        if ws.cell(row=row, column=SETTLEMENT_LABEL_COLUMN).value == label:
            return row
    return None


def _period_label(ws) -> str | None:
    """Řádek s obdobím 'DD.MM.RRRR - DD.MM.RRRR' – jediný text s pomlčkou
    a číslicí v hlavičce souhrnu, hledá se stejně odolně jako popisky."""
    for row in range(1, SETTLEMENT_SEARCH_ROWS + 1):
        value = ws.cell(row=row, column=SETTLEMENT_LABEL_COLUMN).value
        if isinstance(value, str) and "-" in value and any(ch.isdigit() for ch in value):
            return value
    return None


def _parse_settlement(ws, ean: str | None) -> dict[str, Any]:
    """Měsíční souhrn (sloupce K–N). Chybějící řádek necháváme None – radši
    přiznaná neznalost než dopočítaná hodnota ze sumy intervalů, která by
    nemusela sedět s reálnou fakturací (zaokrouhlení, poplatky)."""

    def amount_kwh(key: str) -> float | None:
        row = _find_label_row(ws, SETTLEMENT_LABELS[key])
        if row is None:
            return None
        value = ws.cell(row=row, column=SETTLEMENT_AMOUNT_COLUMN).value
        return None if value is None else round(float(value) * 1000, 3)   # MWh -> kWh

    def total_czk(key: str) -> float | None:
        row = _find_label_row(ws, SETTLEMENT_LABELS[key])
        if row is None:
            return None
        value = ws.cell(row=row, column=SETTLEMENT_TOTAL_COLUMN).value
        return None if value in (None, "-") else float(value)

    return {
        "ean": ean,
        "period_label": _period_label(ws),
        "delivered_kwh": amount_kwh("delivered_kwh"),
        "shared_kwh": amount_kwh("shared_kwh"),
        "sold_kwh": amount_kwh("sold_kwh"),
        "deduction_czk": total_czk("deduction_czk"),
        "invoice_czk": total_czk("invoice_czk"),
    }


def parse_sharing_export(path: Path, *, load_workbook=None) -> dict[str, Any]:
    """Naparsuje jeden měsíční export: intervalová data + měsíční souhrn.

    `load_workbook` je injektovatelné (stejný vzor jako `session`/`connect`
    u ostatních readerů) – testy tak nepotřebují mít nainstalovaný openpyxl,
    stačí jim malá falešná sešitová struktura.
    """
    wb = (load_workbook or _load_workbook)(path)
    ws = wb[wb.sheetnames[0]]
    columns = _header_columns(ws)
    intervals = _parse_intervals(ws, columns)
    if not intervals:
        raise ValueError("export neobsahuje žádné intervalové záznamy")

    ean = intervals[-1]["ean"]
    return {
        "ean": ean,
        "period_from": intervals[0]["timestamp"].date(),
        "period_to": intervals[-1]["timestamp"].date(),
        "intervals": intervals,
        "settlement": _parse_settlement(ws, ean),
    }
